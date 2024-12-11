from django.db.models import Q 
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from myapp.forms import *
from myapp.models import *
from datetime import datetime
from .utils import parse_date
import json
import pandas as pd
import xlsxwriter
import openpyxl
import os
from openpyxl.styles import Font




# Create your views here.

LOGIN_URL = 'main:login'

def login_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('main:index')
        else:
            return redirect('main:index')
    return render(request, 'login.html')

@login_required(login_url='main:login')
def logout_user(request):
    logout(request)
    return redirect('main:login')




@login_required(login_url='main:login')
def index(request):
    times = datetime.now()
    userNote = UsersNote.objects.all()
    context = {
        'userNote':userNote,
        'times':times
    }
    return render(request,'index.html', context)



@login_required(login_url='main:login')
def table(request):
    times = datetime.now()
    user_devices = UserDevice.objects.all()
 
    if 'export_excel' in request.GET:
        data = []
        for devise in user_devices:
            imeis = devise.imeis.all()  
            imei_list = [imei.imei for imei in imeis]
            imei_str = ", ".join(imei_list)

            data.append({
                "Ism": devise.name,
                "Viloyat": devise.categoryRegion.title,
                "Tuman": devise.categoryDistrict.title,
                "Qurilma Turi": devise.device_type.title,
                "Qurilma Nomi": devise.device_name,
                "Yo‘qolgan Vaqti": devise.data.strftime("%d-%m-%Y %H:%M"),
                "Telefon Raqami": devise.phone_number,
                "IMEI": imei_str if imei_str else "N/A",
                "Ariza Fayli": "True" if devise.file else "N/A",
            })

        df = pd.DataFrame(data)

        # Excel faylini yaratish
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="qurilmalar.xlsx"'

      
        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Qurilmalar')

        return response

    context = {
        'times': times,
        'user_devices': user_devices,
    }

    return render(request, 'table.html', context)

def delete_table(request, pk):
    try:
        userDevice = get_object_or_404(UserDevice, pk=pk)
        userDevice.delete()
        messages.success(request, f"Qurilma ID: {pk} muvaffaqiyatli o'chirildi.")
    except Exception as e:
        messages.error(request, f"Qurilmani o'chirishda xatolik yuz berdi: {str(e)}")
    return redirect('main:table')

    
@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def users(request):
    user_objects = Account.objects.all()
    query = request.GET.get('search')
    times = datetime.now()
    
    if query:
        user_objects = user_objects.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(phone_number__icontains=query)
        )
    
    if request.method == 'POST':
        form = AccountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Foydalanuvchi muvaffaqiyatli qo'shildi!")
            return redirect('main:users')
        else:
            messages.error(request, "Formada xatoliklar mavjud. Iltimos, tekshirib qaytadan kiriting.")
    else:
        form = AccountForm()
        
    context = {
        'times': times,
        'users': user_objects,
        'form': form,
    }
    return render(request, 'users.html', context)


@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def update_user(request, pk):
    user = get_object_or_404(Account, pk=pk)
    times = datetime.now()
    if request.method == 'POST':
        form = AccountForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Foydalanuvchi muvaffaqiyatli tahrirlandi.")
            return redirect('main:users')
        else:
            messages.error(request, "Formada xatoliklar mavjud. Iltimos, tekshirib qaytadan kiriting.")
    else:
        form = AccountForm(instance=user)
    context = {'form': form,'times': times,'user': user}
    return render(request, 'create/update_user.html', context)
    
    
    
@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)      
def delete_user(request, pk):
    user = get_object_or_404(Account, pk=pk)
    user.delete()
    return redirect('main:users')    

    
    


@login_required(login_url='main:login')
def statistika(request):
    times = datetime.now()
    user_device_count = UserDevice.objects.count()
    found_device_count = FoundDevice.objects.count()
    statistika = DeviceStatistics.objects.all()
    
    print(found_device_count)
    
    context = {
        'times': times,
        'statistika':statistika,
        'user_device_count':user_device_count,
        'found_device_count':found_device_count,
    }

    return render(request, 'statistika.html', context)

    
@login_required(login_url='main:login')
def qurilma(request):
    times = datetime.now()
    
    if request.method == 'POST':
        user_device_form = UserDeviseForm(request.POST, request.FILES)
        imei_formset = DeviceIMEIFormSet(request.POST)
        
        if user_device_form.is_valid() and imei_formset.is_valid():
            user_device = user_device_form.save(commit=False)
            user_device.save()
            
            for form in imei_formset:
                imei = form.cleaned_data.get('imei')
                if imei:
                    DeviceImei.objects.create(device=user_device, imei=imei)
            
            return redirect('main:table')
        else:
            return render(request, 'qurilma.html', {
                'user_device_form': user_device_form,
                'imei_formset': imei_formset,
                'times': times,
            })
    else:
        user_device_form = UserDeviseForm()
        imei_formset = DeviceIMEIFormSet()
        return render(request, 'qurilma.html', {
            'user_device_form': user_device_form,
            'imei_formset': imei_formset,
            'times': times,
        })



@user_passes_test(lambda user: user.is_staff)  
@login_required(login_url='main:login')
def update_qurilma(request, pk):
    times = datetime.now()
    qurilma = get_object_or_404(UserDevice, pk=pk)
    
    if request.method == 'POST':
        user_device_form = UserDeviseForm(request.POST, request.FILES, instance=qurilma)
        imei_formset = DeviceIMEIFormSet(request.POST)
        if user_device_form.is_valid() and imei_formset.is_valid():
            user_device = user_device_form.save()
            
            DeviceImei.objects.filter(device=user_device).delete()
            
            for form in imei_formset:
                imei = form.cleaned_data.get('imei')
                if imei:
                    DeviceImei.objects.create(device=user_device, imei=imei)
            
            return redirect('main:table')
        else:
            return render(request, 'update_qurilma.html', {
                'form': user_device_form,
                'imei_formset': imei_formset,
                'qurilma': qurilma,
                'times': times,
            })
    
    else:
        
        user_device_form = UserDeviseForm(instance=qurilma)
        imei_formset = DeviceIMEIFormSet(queryset=DeviceImei.objects.filter(device=qurilma))  # Fetch related IMEIs

        return render(request, 'update_qurilma.html', {
            'form': user_device_form,
            'imei_formset': imei_formset,
            'qurilma': qurilma,
            'times': times,
        })






@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def setting(request):
    times = datetime.now()
    context = {
        'times':times
    }
    return render(request, 'create/base.html', context)


@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def add_region(request):
    region = Region.objects.all()
    times = datetime.now()
    if request.method == 'POST':
        form = RegionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('main:add_region')
    else:
        form = RegionForm()
    return render(request, 'create/add_region.html', {
        'form': form,
        'times': times,
        'region':region
        })



@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def add_device_type(request):
    device = DeviceType.objects.all()
    times = datetime.now()
    if request.method == 'POST':
        form = DeviceTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('main:add_device_type')
    else:
        form = DeviceTypeForm()
    return render(request, 'create/add_device_type.html', {'form': form,'times': times,'device': device})



@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def add_district(request):
    district = District.objects.all()  
    times = datetime.now()
    if request.method == 'POST':
        form = DistrictForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('main:add_district')
    else:
        form = DistrictForm()
    
    return render(request, 'create/add_district.html', {'form': form, 'times': times, 'district': district})








@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def update_district(request, pk):
    district = get_object_or_404(District, pk=pk)
    times = datetime.now()
    
    if request.method == 'POST':
        form = DistrictForm(request.POST, instance=district)
        if form.is_valid():
            form.save()
            return redirect('main:add_district')
    else:
        form = DistrictForm(instance=district)
    
    return render(request, 'create/update_district.html', {'form': form, 'times': times, 'district': district})



@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def add_userNote(request):
    userNote = UsersNote.objects.all()
    times = datetime.now()
    if request.method == 'POST':
        form = UsersNoteForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('main:add_note')
    else:
        form = UsersNoteForm()
    
    return render(request, 'create/add_note.html', {'form': form, 'times': times,'userNote':userNote})



@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def update_userNote(request, pk):
    userNote = get_object_or_404(UsersNote, pk=pk)
    times = datetime.now()
    
    if request.method == 'POST':
        form = UsersNoteForm(request.POST, instance=userNote)
        if form.is_valid():
            form.save()
            return redirect('main:add_note')
    else:
        form = UsersNoteForm(instance=userNote)
    
    return render(request, 'create/update_note.html', {'form': form, 'times': times, 'userNote': userNote})



@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def delete_note(request, pk):
    userNote = get_object_or_404(UsersNote, pk=pk)
    userNote.delete()
    return redirect('main:add_note') 




@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def delete_district(request, pk):
    district = get_object_or_404(District, pk=pk)
    district.delete()
    return redirect('main:add_district') 



@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def update_device(request, pk):
    device = get_object_or_404(DeviceType, pk=pk)
    times = datetime.now()
    
    if request.method == 'POST':
        form = DeviceTypeForm(request.POST, instance=device)
        if form.is_valid():
            form.save()
            return redirect('main:add_device_type')
    else:
        form = DeviceTypeForm(instance=device)
    return render(request, 'create/update_device.html', {
        'form': form,
        'times': times,
        'device': device
    })
 


@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def delete_device(request, pk):
    device = get_object_or_404(DeviceType, pk=pk)
    device.delete()
    return redirect('main:add_device_type')



@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)
def update_region(request, pk):
    region = get_object_or_404(Region, pk=pk)
    times = datetime.now()
    if request.method == 'POST':
        form = RegionForm(request.POST, instance=region)
        if form.is_valid():
            form.save()
            return redirect('main:add_region')
    else:
        form = RegionForm(instance=region)
        return render(request, 'create/update_region.html', {
        'form': form,
        'times': times,
        'region': region
    })
        
        
@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_superuser)      
def delete_region(request, pk):
    region = get_object_or_404(Region, pk=pk)
    region.delete()
    return redirect('main:add_region')





@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_staff)
def found_device(request):
    times = datetime.now()
    file_error = None

    if request.method == 'POST':

        uploaded_file = request.FILES.get('uploaded_file')
        if not uploaded_file:
            file_error = "Faylni yuklang!"
        else:
            file_extension = os.path.splitext(uploaded_file.name)[1].lower()
            if file_extension != '.xlsx':
                file_error = "Fayl faqat .xlsx formatida bo'lishi kerak!"
            else:
                try:
                    wb = openpyxl.load_workbook(uploaded_file)
                    sheet = wb.active

                    for index, row in enumerate(sheet.iter_rows(values_only=True)):
                        row = [item for item in row if item]
                        if index <= 5 or not row or len(row) != 17:
                            continue
                        imei = str(row[5]).strip()
                        operator = row[11]
                        phone_number = str(row[-3]).strip()
                        active_started = parse_date(row[7])
                        active_finished = parse_date(row[8])
                        
                        try:
                            imei_record = DeviceImei.objects.get(imei=imei)
                            if FoundDevice.objects.filter(imei=imei_record, active_started=active_started, active_finished=active_finished).exists():
                                continue
                        except DeviceImei.DoesNotExist:
                            continue
                        
                        FoundDevice.objects.create(
                            imei=imei_record,
                            active_started=active_started,
                            active_finished=active_finished,
                            operator=operator,
                            phone_number=phone_number,
                        )
                except Exception as e:
                    file_error = f"Faylni o'qishda xato!"
                    print(e)


    if 'export' in request.GET:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Found Devices"

        headers = [
            "ID", "IMEI", "Operator", "Phone Number",
            "Active Started", "Active Finished", "Is Viewed"
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        found_devices = FoundDevice.objects.all()
        for device in found_devices:
            ws.append([
                device.id,
                device.imei.imei if device.imei else "",
                device.operator,
                device.phone_number,
                device.active_started.strftime('%Y-%m-%d') if device.active_started else "",
                device.active_finished.strftime('%Y-%m-%d') if device.active_finished else "",
                "Yes" if device.is_viewed else "No"
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="found_devices.xlsx"'
        wb.save(response)
        return response

    context = {
        'times': times,
        'file_error': file_error,
        'found_devices': FoundDevice.objects.all(),
    }
    return render(request, 'found.html', context)


@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_staff)
def found_device_detail(request, pk):
    times = datetime.now()
    device = get_object_or_404(FoundDevice, pk=pk)
    device.is_viewed = False
    device.save()
    return render(request, 'found-detail.html', {'device': device,'times': times,})



@login_required(login_url='main:login')
@user_passes_test(lambda user: user.is_staff)
def delete_found_device(request, pk):
    device = get_object_or_404(FoundDevice, pk=pk)
    device.delete()
    return redirect('main:found')